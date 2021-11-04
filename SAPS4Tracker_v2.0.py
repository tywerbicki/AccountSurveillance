depositPath = '$$$$$$$$$$'
dependenciesDir = '$$$$$$$$$$'

from os import chmod, getcwd, path, mkdir, remove, listdir
from stat import S_IWUSR, S_IWOTH, S_IRUSR, S_IROTH
from time import sleep
from msvcrt import kbhit
from sys import exit
from traceback import print_exc
from datetime import date, timedelta, datetime, timezone
from pytz import timezone as pytzTZ
from pandas import read_excel, DataFrame, concat, ExcelWriter
from numpy import empty, empty_like, full, zeros, object, float64, int64, abs, array
from numpy import sum as np_sum
from win32com.client import Dispatch as win32_Dispatch
from openpyxl import load_workbook
from openpyxl.styles import Alignment

#Function Definitions
def AcquireProgramLock():
    '''
    Acquires the program lock so that another user cannot update the database simultaneously.
    '''
    chmod(dependenciesDir + '$$$$$$$$$$', S_IWUSR | S_IWOTH) 
    try:
        with open(dependenciesDir + '$$$$$$$$$$', 'r+') as f:
            text = f.read()
            if text == '$$$$$$$$$$':
                f.seek(0)
                f.write('$$$$$$$$$$') 
                f.truncate()
            elif text == '$$$$$$$$$$':
                print('Another User is Currently Running This Program. Please Try Again Shortly.')
                sleep(3)
                print("Program Exiting Safely.")
                sleep(2)
                exit()
    finally:
        chmod(dependenciesDir + '$$$$$$$$$$', S_IRUSR | S_IROTH) #make file read only

def ReleaseProgramLock():
    '''
    Releases the program lock when finished so that another user can access the program.
    '''
    chmod(dependenciesDir + '$$$$$$$$$$', S_IWUSR | S_IWOTH) 
    with open(dependenciesDir + '$$$$$$$$$$', 'w') as f:
        f.write('$$$$$$$$$$') 
    chmod(dependenciesDir + '$$$$$$$$$$', S_IRUSR | S_IROTH) 
    return

def AcquireDates(numDelta = 0):
    '''
    Acquires the dates for the program to use when searching for emails and documents.
    Accounts for business days by excluding weekends.
    The default argument is to fetch the appropriate dates for the current day, but the user can go back
    any number of days via the timedelta implementation.
    '''
    rawDate = date.today()
    currentDate = (rawDate - timedelta(numDelta)).strftime("%m/%d/%Y")
    if (rawDate - timedelta(numDelta)).weekday() < 2: 
        propRecvDate_DB = (rawDate - timedelta(4 + numDelta)).strftime("%m/%d/%Y") 
    else:
        propRecvDate_DB = (rawDate - timedelta(2 + numDelta)).strftime("%m/%d/%Y")

    if (rawDate - timedelta(numDelta)).weekday() == 0:
        priorBusinessDay = (rawDate - timedelta(3 + numDelta)).strftime("%m/%d/%Y") 
    else:
        priorBusinessDay = (rawDate - timedelta(1 + numDelta)).strftime("%m/%d/%Y")

    if (rawDate - timedelta(numDelta)).weekday() == 4:
        propFocDate = (rawDate + timedelta(3 - numDelta)).strftime("%m/%d/%Y") 
    else:
        propFocDate = (rawDate + timedelta(1 - numDelta)).strftime("%m/%d/%Y")

    return propFocDate, currentDate, priorBusinessDay, propRecvDate_DB

def UpdateInbox():
    '''
    Moves emails from inbox into appropriate folders prior to the program running.
    THIS FUNCTION WILL EXECUTE EVERY TIME THE PROGRAM RUNS.
    '''

    def AnalyzeSAPS4Email(message):
        analyzeMessage = False
        if message.Class == 43:
            if message.SenderEmailType == 'EX':
                if message.Sender.GetExchangeUser().PrimarySmtpAddress == '$$$$$$$$$$': 
                    analyzeMessage = True
        return analyzeMessage

    def AnalyzeTreasuraEmail(message):
        analyzeMessage = False
        if message.Class == 43:
            if message.SenderEmailType == 'SMTP':
                if isinstance(message.Sender.GetExchangeUser(), type(None)):
                    if '$$$$$$$$$$' in message.subject.lower().replace(' ', ''):
                        analyzeMessage = True
        return analyzeMessage

    def MoveProposedOutflowEmail(message):
        '''
        Returns whether or not an email should be moved into the Proposed Daily Outflows folder.
        These emails to be moved are sent by this very program, so this is more of a housekeeping
        functionality to prevent these proposed outflow emails from cluttering up one's inbox.
        '''
        moveMessage = False
        if message.Class == 43:
            if message.SenderEmailType == 'EX':
                if '$$$$$$$$$$' in message.subject:
                    moveMessage = True
        return moveMessage

    outlook = win32_Dispatch('outlook.application').GetNamespace('MAPI')
    inbox = outlook.GetDefaultFolder(6)

    inboxFolders = {folder.Name for folder in inbox.Folders}
    if '$$$$$$$$$$' in inboxFolders:
        mainFolder = inbox.Folders.Item('$$$$$$$$$$')
    else:
        inbox.Folders.Add('$$$$$$$$$$')
        mainFolder = inbox.Folders.Item('$$$$$$$$$$')
        print("Created Folder In Inbox Titled: '$$$$$$$$$$'")

    mainFolders = {folder.Name for folder in mainFolder.Folders}
    if '$$$$$$$$$$' in mainFolders:
        SAP_Folder = mainFolder.Folders.Item('$$$$$$$$$$')
    else:
        mainFolder.Folders.Add('$$$$$$$$$$')
        SAP_Folder = mainFolder.Folders.Item('$$$$$$$$$$')
        print("Created Folder In '$$$$$$$$$$' Titled: '$$$$$$$$$$'")
    if '$$$$$$$$$$' in mainFolders:
        Treas_Folder = mainFolder.Folders.Item('$$$$$$$$$$')
    else:
        mainFolder.Folders.Add('$$$$$$$$$$')
        Treas_Folder = mainFolder.Folders.Item('$$$$$$$$$$')
        print("Created Folder In '$$$$$$$$$$' Titled: '$$$$$$$$$$'")
    if '$$$$$$$$$$' in mainFolders:
        proposedOutflowFolder = mainFolder.Folders.Item('$$$$$$$$$$')
    else:
        mainFolder.Folders.Add('$$$$$$$$$$')
        proposedOutflowFolder = mainFolder.Folders.Item('$$$$$$$$$$')
        print("Created Folder In '$$$$$$$$$$' Titled: '$$$$$$$$$$'")

    inbmessages = inbox.Items
    inbmessages.sort("[ReceivedTime]", True) 
    inbmessage = inbmessages.GetFirst()
    nonTargetEmails = 0
    emailCounter = 0

    while emailCounter < 200: 
        if not isinstance(inbmessage, type(None)):
            emailCounter += 1
            analyzedEmail = False

            if AnalyzeSAPS4Email(inbmessage):
                inbmessage.Unread = True ; analyzedEmail = True
                inbmessage.Move(SAP_Folder)
                print(f"MOVED EMAIL ({inbmessage.subject}) to '$$$$$$$$$$'")

            elif AnalyzeTreasuraEmail(inbmessage):
                inbmessage.Unread = True ; analyzedEmail = True
                inbmessage.Move(Treas_Folder)
                print(f"MOVED EMAIL ({inbmessage.subject}) to '$$$$$$$$$$'")

            elif MoveProposedOutflowEmail(inbmessage):
                inbmessage.Unread = True ; analyzedEmail = True
                inbmessage.Move(proposedOutflowFolder)
                print(f"MOVED EMAIL ({inbmessage.subject}) to '$$$$$$$$$$'")

            if analyzedEmail:
                inbmessages = inbox.Items 
                inbmessages.sort("[ReceivedTime]", True)
                inbmessage = inbmessages.GetFirst()
                for _ in range(nonTargetEmails):
                    inbmessage = inbmessages.GetNext()
            else:
                nonTargetEmails += 1
                inbmessage = inbmessages.GetNext()

        else:
            break

def InitializeApplication():
    '''
    Acquires the date from the user. Allows the user to run the program for a previous day in
    case they were away. Prevents the user from updating the database on a day residing on the
    weekend (as there are no proposals/registers sent on the weekend), as well as updating the
    database earlier than 13:00 eastern time, to make sure that all of the day's proposals have
    been sent. Also, prevents the database from being updated on the same day twice, by referencing
    the databaseupdatesdates.txt file.
    '''
    day = 'unknown'
    while day == 'unknown':

        secondsUntilProceed = 10 ; pollingRate = .1
        day = "same"
        print("Please Press Any Key To Change The Date. You have 10 Seconds To Do So.")
        while secondsUntilProceed > 0:
            if abs(secondsUntilProceed % 2) < 0.0001:
                print(f"{int(secondsUntilProceed)}s ...", end = " ", flush = True)
            sleep(pollingRate)
            secondsUntilProceed -= pollingRate
            if kbhit():
                day = "change"
                break
        print()

        if day == 'same':
            if date.today().weekday() > 4:
                print('The Database Cannot Be Updated For A Day Residing On The Weekend.')
                sleep(4)
                print('Returning To Start-Up Menu.')
                sleep(2)
                day = 'unknown'
                continue
            if datetime.now(timezone.utc).astimezone(pytzTZ('Canada/Eastern')).hour < 13:
                print('The Database Cannot Be Updated Sooner Than 13:00 Eastern Time For The Current Day.')
                sleep(4)
                print('Program Exiting Safely.')
                sleep(2)
                exit()
            with open(dependenciesDir + '$$$$$$$$$$', 'r') as f:
                dates = f.read()
                datesList = dates.split(';')
                if date.today().strftime("%m/%d/%Y") in datesList:
                    print("The Database Has Already Been Updated For", date.today().strftime("%m/%d/%Y"))
                    sleep(4)
                    print('Program Exiting Safely.')
                    sleep(2)
                    exit()

            return AcquireDates()

        elif day == 'change':
            newday = input("Please Enter an Integer For the Number of Days Prior To Analyze (>= 1): ")
            try:
                newday = int(newday)
            except ValueError:
                print(f"{newday} Is Not An Integer. Returning To Start-Up Menu.")
                day = 'unknown'
                sleep(3)
                continue

            if (date.today() - timedelta(newday)).weekday() > 4:
                print('The Database Cannot Be Updated For A Day Residing On A Weekend.')
                sleep(4)
                print('Returning To Start-Up Menu.')
                sleep(2)
                day = 'unknown'
                continue

            with open(dependenciesDir + '$$$$$$$$$$', 'r') as f:
                dates = f.read()
                datesList = dates.split(';')
                if (date.today() - timedelta(newday)).strftime("%m/%d/%Y") in datesList:
                    print("The Database Has Already Been Updated For", (date.today() - timedelta(newday)).strftime("%m/%d/%Y"))
                    sleep(4)
                    print('Program Exiting Safely.')
                    sleep(2)
                    exit()

            return AcquireDates(newday) 

def PropRegTreasFolderCleanup(currentDate, propDir, regDir, treasDir):
    '''
    Removes files older than 31 days in the directories where the proposals, registers,
    and AP EFT reports are being deposited.
    Prevents hundreds of files accumulating in these directories and taking up large amounts
    of space on the corporate drive.
    '''
    if datetime.strptime(currentDate, "%m/%d/%Y").weekday() > 0: 
        return

    def FolderCleaner(directory, fileNames):
        for file in fileNames:
            filePath = directory + '/' + file
            fileDate = datetime.fromtimestamp(path.getmtime(filePath))
            if (datetime.today() - fileDate).total_seconds() > 60*60*24*31: 
                remove(filePath) 

    propFileNames = listdir(propDir) 
    regFileNames = listdir(regDir)
    treasFileNames = listdir(treasDir)
    FolderCleaner(propDir, propFileNames)
    FolderCleaner(regDir, regFileNames)
    FolderCleaner(treasDir, treasFileNames)

def AcquireFilesFromOutlook(priorBusinessDay, currentDate, regDir, treasDir):
    '''
    Downloads the appropriate payment registers and treasura AP EFT report from the outlook client onto
    the corporate drive in the 'regDir' and 'treasDir' directories, respectively.
    '''
    def HasDate(string):
        hits = [c.isdigit() for c in string]
        return True if sum(hits) > 4 else False

    outlook = win32_Dispatch('outlook.application').GetNamespace('MAPI')
    inbox = outlook.GetDefaultFolder(6)
    SAP_Folder = inbox.Folders.Item('$$$$$$$$$$').Folders.Item('$$$$$$$$$$')
    SAP_messages = SAP_Folder.Items
    SAP_messages.sort("[ReceivedTime]", True)
    Treas_Folder = inbox.Folders.Item('$$$$$$$$$$').Folders.Item('$$$$$$$$$$')
    Treas_messages = Treas_Folder.Items
    Treas_messages.sort("[ReceivedTime]", True)
    proposedOutflowFolder = inbox.Folders.Item('$$$$$$$$$$').Folders.Item('$$$$$$$$$$')
    proposedOutflow_messages = proposedOutflowFolder.Items
    proposedOutflow_messages.sort("[ReceivedTime]", True)

    SAP_message = SAP_messages.GetFirst()
    for _ in range(20):
        if not isinstance(SAP_message, type(None)):
            if ('$$$$$$$$$$' in SAP_message.subject.lower()) or ('$$$$$$$$$$' in SAP_message.subject.lower()):
                pass
            elif '$$$$$$$$$$' and '$$$$$$$$$$' in SAP_message.subject.lower().replace(' ', ''):
                if currentDate == SAP_message.SentOn.strftime("%m/%d/%Y"):
                    for attachment in SAP_message.Attachments:
                        if '.xls' in attachment.filename and HasDate(attachment.filename):
                            attachment.SaveAsFile(propDir + '\\' + attachment.filename)
                            print(f"Downloaded File:  {attachment.filename}")
            elif '$$$$$$$$$$' in SAP_message.subject.lower().replace(' ', '') and priorBusinessDay == SAP_message.SentOn.strftime("%m/%d/%Y"):
                for attachment in SAP_message.Attachments:
                    if '$$$$$$$$$$' and '.xls' in attachment.filename.lower() and HasDate(attachment.filename):
                        attachment.SaveAsFile(regDir + '\\' + attachment.filename)
                        print(f"Downloaded File:  {attachment.filename}")

            SAP_message = SAP_messages.GetNext()
        else:
            break

    outflowReportAcquired = False
    Treas_message = Treas_messages.GetFirst()
    for _ in range(5):
        if not isinstance(Treas_message, type(None)):
            if currentDate == Treas_message.SentOn.strftime("%m/%d/%Y"):
                if '$$$$$$$$$$' in Treas_message.body.lower().replace(' ', ''):
                    for attachment in Treas_message.Attachments:
                        if '$$$$$$$$$$' and '.xls' in attachment.filename.lower():
                            fileName = attachment.filename[:-4] + '_' + priorBusinessDay.replace('/', "") + '.xls'
                            attachment.SaveAsFile(treasDir + '\\' + fileName)
                            outflowReportAcquired = True
                            print(f"Downloaded File:  {fileName}")
                            print()
                            break

            if outflowReportAcquired:
                break
            else:
                Treas_message = Treas_messages.GetNext()
        else:
            break

def SAPFileDetector(directory, date, reportType):
    '''
    Searches a given directory (proposal or register) and returns a list with each of the file names to analyze.
    '''

    def CheckDate (date, fileToBeChecked):
        '''
        Converts a date string mm/dd/yyyy into mmddyy to match what is found in the filenames sent from SAP S4.
        These filenames on the SAP S4 files can be used to identify what day that file is applicable to.
        '''
        toks = date.split('/')
        dateMod = toks[0] + toks[1] + toks[2][-2:]
        return True if dateMod in fileToBeChecked else False

    files_needed = []
    for file in directory:
        if CheckDate(date, file):
            files_needed.append(file)
    if len(files_needed) > 0:
        return files_needed
    else: #sometimes there will be no files for analysis. This could be the result of the proposals not being sent yet, or that the next business day is a holiday.
        cont = input(f"No {reportType}s Found For {date}.  Do You Wish To Continue? (y/n): ") #want to continue if no proposals due to holiday
        if cont == 'y':
            return files_needed
        elif cont == 'n': #do not want to continue if the proposals have yet to be sent. Rerun the program manually after the proposals have been sent.
            print('Program Exiting Safely.')
            sleep(2)
            exit()
        else:
            print('Incorrect Entry.')
            sleep(2)
            print('Program Exiting Safely.')
            sleep(2)
            exit()

def TreasuraFileDetector(treasDir, treasFilenames, currentDate):
    '''
    Searches the treasura file directory for the filename of the report to analyze.
    '''
    def CheckDate (date, directory, fileName):
        fileDate = datetime.fromtimestamp(path.getmtime(directory + '\\' + fileName)).strftime("%m/%d/%Y")
        return True if date == fileDate else False 

    fileNeeded = None
    for file in treasFilenames:
        if CheckDate(currentDate, treasDir, file):
            fileNeeded = file
    if fileNeeded == None:
        raise FileNotFoundError('Please Wait for Incoming Data from Treasura') 
    else:
        return treasDir + '\\' + fileNeeded

def NameStripper(colnames):
    '''
    Takes in the column names list returned by the above function. Removes the whitespace from
    each label. This is done because sometimes the SAP S4 reports have random spaces in their
    column names, making them hard to work with during downstream analysis.
    '''
    return [name.strip() if isinstance(name, str) else 'Unknown' for name in colnames]

def CompanyFilter(dataframe, companyCode, currency, TYPE = 'prop'):
    """
    Returns index that fetches rows of the desired company, and the currency of the transactions of that
    company. This is important, because some of the companies we are looking at have CAD and USD branches,
    whos transactions are concerned with different Nutrien account structures.
    """
    if TYPE == 'prop':
        index = array(dataframe['$$$$$$$$$$'] == companyCode) & array(dataframe['$$$$$$$$$$'] == currency)
    elif TYPE == 'reg':
        index = array(dataframe['$$$$$$$$$$'] == companyCode) & array(dataframe['$$$$$$$$$$'] == currency)
    return index

def PaymentErrorFilter(dataframe):
    """
    Returns a boolean index in list form that fetches all of the columns associated with True in
    the boolean index. In this case, the function is creating an index that will retain only payments
    with NO error message. In the database and the proposed outflows file, this filtering process
    generates the column tagged with NO_ERRORS.
    """
    return dataframe['$$$$$$$$$$'].isna()

def ToVisualFormat(num):
    '''
    Makes currency values easier to look at.
    '''
    value = str(abs(round(num, 2)))
    if "." in value:
        body, deci = value.split('.')
        deci = "." + deci
    else:
        body = value ; deci = ""

    slicer = 3
    while slicer:
        if len(body) > slicer:
            body = body[:-slicer] + "," + body[-slicer:]
            slicer += 4
        else:
            slicer = False

    if num < 0:
        return "(" + body + deci + ")"
    else:
        return body + deci

def DataAgreggator_F(companyList, currencyList, presentDayProposals, propDir, propFocDate):
    '''
    Aggregates all of the appropriate data in the proposals and combines it into an output that
    can be used for the proposed outflows email.
    THIS FUNCTION IS CONCERNED WITH THE CURRENT DATE PROPOSALS, WHICH ARE PROPOSING PAYMENTS FOR
    THE FOLLOWING BUSINESS DAY.
    Then updates the supplementary data to be referenced by the morning account info spreadsheet.
    Concludes by returning the names of the proposals analyzed so that they can be included on the
    email sent out to analysts.
    '''

    def PropDataScraper_F(companyList, currencyList, presentDayProposals, propDir, propFocDate):
        pooledPropOutflows = zeros(3, float64)
        analyzedNames = ""
        for file in presentDayProposals:
            try:
                data = read_excel(propDir + "\\" + file)
                data.columns = NameStripper(data.columns) 
                counter = 0
                for companyCode, currency in zip(companyList, currencyList):
                    tempData = data[CompanyFilter(data, companyCode, currency)]
                    if len(tempData) > 0:
                        if counter < 2:
                            index = 0
                        elif counter < 5:
                            index = 1
                        else:
                            index = 2
                        pooledPropOutflows[index] += tempData[PaymentErrorFilter(tempData)]['$$$$$$$$$$'].sum()
                    counter += 1

                analyzedNames += '\t' + file + '\n' 

            except Exception as e:
                print(f"ERROR analyzing {file}")
                print(e)
                analyzedNames += '\t' + 'ERROR analyzing: ' + file + '\n'

        return pooledPropOutflows, analyzedNames

    pooledPropOutflows, analyzedNames = PropDataScraper_F(companyList, currencyList, presentDayProposals, propDir, propFocDate) 
    structures = ['$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$']
    messageForEmail = "\n"
    for struc, pool in zip(structures, pooledPropOutflows):
        messageForEmail += '\t' + struc + ": " + ToVisualFormat(pool) + '\n\n'

    chmod(depositPath, S_IWUSR | S_IWOTH)
    DATABASE_WORKBOOK = load_workbook(depositPath)
    supDataSheet = DATABASE_WORKBOOK['$$$$$$$$$$']
    titleCell = supDataSheet.cell(row = 1, column = 1)
    title, date = titleCell.value.split(":")
    titleCell.value = title + ":  " + propFocDate
    amountCell = supDataSheet.cell(row = 1, column = 2)
    amountCell.value = pooledPropOutflows[1]
    DATABASE_WORKBOOK.save(depositPath)
    chmod(depositPath, S_IRUSR | S_IROTH)

    return messageForEmail, analyzedNames

def DataAgreggator_DB(companyList, currencyList, accountList, propFlaggedFiles, propDir, propRecvDate, regFlaggedFiles, regDir, regRecvDate, treasFile):
    '''
    Aggregates all of the appropriate data in the proposals and registers and combines it into an output that
    can be used to update the database.
    THIS FUNCTION IS CONCERNED WITH THE PREVIOUS PREVIOUS BUSINESS DAY'S PROPOSALS, so that these can be compared
    with the previous business day's register and treasura AP EFT report (this report is for the previous business
    day but SENT ON THE CURRENT DAY).
    '''
    def PropDataScraper_DB(companyList, currencyList, propFlaggedFiles, propDir, propRecvDate):
        dataForDataBase = empty([8, 9], dtype = object)
        dataForDataBase[:, 2] = full((8,), propRecvDate, dtype = object)
        dataForDataBase[:, 4] = companyList
        dataForDataBase[:, 5] = currencyList
        dataForDataBase[:, [6, 7, 8]] = zeros((8, 3), dtype = float64)

        for file in propFlaggedFiles:
            try:
                data = read_excel(propDir + '/' + file)
                data.columns = NameStripper(data.columns)
                counter = 0
                for companyCode, currency in zip(companyList, currencyList):
                    tempData = data[CompanyFilter(data, companyCode, currency)]
                    if len(tempData) > 0:
                        if dataForDataBase[counter, 0] == None:
                            dataForDataBase[counter, 0] = file
                        else:
                            dataForDataBase[counter, 0] += " | " + file
                        if 'EFT' in file or 'ACH' in file or 'FRTCE' in file or 'FRTUA' in file:
                            dataForDataBase[counter, 6] += tempData[PaymentErrorFilter(tempData)]['Net Amount in FC'].sum()
                        elif 'WIRE' in file:
                            dataForDataBase[counter, 7] += tempData[PaymentErrorFilter(tempData)]['Net Amount in FC'].sum()
                        elif 'SCOCA' in file or 'BMOUS' in file or 'FRTCC' in file or 'FRTUC' in file or 'CBRCC' in file:
                            dataForDataBase[counter, 8] += tempData[PaymentErrorFilter(tempData)]['Net Amount in FC'].sum()
                        else:
                            pass
                    else:
                        pass
                    counter += 1

                print(f"Analyzed {file} for Surveillance Database")

            except Exception as e:
                print(f"ERROR analyzing {file} for Surveillance Database")
                print(e)

        return dataForDataBase

    def RegDataScraper(companyList, currencyList, regFlaggedFiles, regDir, regRecvDate):

        def Reg_EFTACHFilter(dataframe):
            '''
            Returns index that fetches only ACH or EFT payments.
            '''
            def Filter(s):
                return 'eft' in s.lower() or 'ach' in s.lower()

            return dataframe['$$$$$$$$$$'].apply(Filter)

        def Reg_CheckFilter(dataframe):
            '''
            Returns index that fetches only check payments.
            '''
            def Filter(s):
                return 'check' in s.lower()

            return dataframe['$$$$$$$$$$'].apply(Filter)

        def Reg_WireFilter(dataframe):
            '''
            Returns index that fetches only wire payments.
            '''
            def Filter(s):
                return 'wire' in s.lower()

            return dataframe['$$$$$$$$$$'].apply(Filter)

        dataForDataBase = empty([8, 5], dtype = object)
        dataForDataBase[:, 1] = full((8,), regRecvDate, dtype = object)
        dataForDataBase[:, [2, 3, 4]] = zeros((8, 3), dtype = float64)

        for file in regFlaggedFiles: #now we are analyzing registers
            try:
                data = read_excel(regDir + '/' + file, sheet_name = 2) #usually data is on second sheet
                if '$$$$$$$$$$' in data.columns:
                    pass
                else:
                    data = read_excel(regDir + '/' + file, sheet_name = 1) 
                    assert '$$$$$$$$$$' in data.columns, "'$$$$$$$$$$' Column Not Found in " + file

                data.columns = NameStripper(data.columns)
                counter = 0
                for companyCode, currency in zip(companyList, currencyList):
                    tempData = data[CompanyFilter(data, companyCode, currency, 'reg')]
                    if len(tempData) > 0:
                        dataForDataBase[counter, 0] = file
                        dataForDataBase[counter, 2] -= tempData[Reg_EFTACHFilter(tempData)]['$$$$$$$$$$'].sum()
                        dataForDataBase[counter, 3] -= tempData[Reg_CheckFilter(tempData)]['$$$$$$$$$$'].sum()
                        dataForDataBase[counter, 4] -= tempData[Reg_WireFilter(tempData)]['$$$$$$$$$$'].sum()
                    counter += 1

                print(f"Analyzed {file} for Surveillance Database")

            except Exception as e:
                print(f"ERROR analyzing {file} for Surveillance Database")
                print(e)

        return dataForDataBase

    def TreasDataScraper(accountList, treasFile, regRecvDate):
        '''
        Analyzes the data contained in the treasura AP EFT reports.
        '''
        dataForDataBase = empty([8, 4], dtype = object)
        dataForDataBase[:, 3] = full((8,), regRecvDate, dtype = object)
        dataForDataBase[:, :3] = zeros([8, 3])

        try:
            treasData = read_excel(treasFile, header = 6)
            treasData = treasData.loc[:, ~treasData.columns.str.contains('Unnamed:')]
            treasData = treasData[~treasData['LEDGER AMOUNT'].isna()]
            treasData["ACCOUNT"] = treasData["ACCOUNT"].astype(int64).astype(str)

            subcategories = ['AP EFT', 'AP Wires', 'AP Cheques']
            for accountIndex, accountNumber in enumerate(accountList):
                tempData1 = treasData[treasData["ACCOUNT"] == accountNumber]
                for subIndex, subcategory in enumerate(subcategories):
                    tempData2 = tempData1[tempData1['SUBCATEGORY'] == subcategory]
                    dataForDataBase[accountIndex, subIndex] += tempData2['LEDGER AMOUNT'].sum()

            fn = treasFile.split('\\')[-1]
            print(f"Analyzed {fn} for Surveillance Database")

        except Exception as e:
            print("ERROR Analyzing the Treasura Report")
            print(e)
            print("Has the Format of the Treasura Report Changed?")

        return dataForDataBase

    finalData = empty([8, 16], dtype = object)
    finalData[:, :9] = PropDataScraper_DB(companyList, currencyList, propFlaggedFiles, propDir, propRecvDate) 
    finalData[:, [1, 3, 9, 10, 11]] = RegDataScraper(companyList, currencyList, regFlaggedFiles, regDir, regRecvDate) 
    finalData[:, 12:] = TreasDataScraper(accountList, treasFile, regRecvDate) 

    return finalData

def ExecuteProposedOutflowsAndDatabaseUpdates(wd, propDir, regDir, treasDir, propFocDate, currentDate, priorBusinessDay, propRecvDate_DB):

    propFilenames = listdir(propDir) 
    regFilenames = listdir(regDir) 
    treasFilenames = listdir(treasDir) 

    companyList = ['$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$'] 
    currencyList = ['USD', 'USD', 'CAD', 'CAD', 'CAD', 'USD', 'USD', 'USD']
    accountList = ['$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$']

    #Daily Outflow Proposals Program Implementation
    presentDayProposals = SAPFileDetector(propFilenames, propFocDate, 'Proposal') 
    messageForEmail, analyzedNames = DataAgreggator_F(companyList, currencyList, presentDayProposals, propDir, propFocDate) 
    print("Daily Proposed Outflows Acquired") ; print()

    names = []
    emailAdresses = []
    with open(dependenciesDir + '$$$$$$$$$$', 'r') as f: 
        namesAndEmails = f.read().split('\n')
        for pair in namesAndEmails:
            name, emailAdress = pair.split(': ')
            names.append(name)
            emailAdresses.append(emailAdress)

    outlook = win32_Dispatch('outlook.application')
    for email, name in zip(emailAdresses, names):
        newEmail = outlook.CreateItem(0) 
        newEmail.Subject = f"Proposed Daily Outflows for {propFocDate}"
        newEmail.Body = f'''Hello {name},

    This is an automated message from Ty's SAP S4 v2.0 Payment Proposal Tracker.


    Please observe the pooled proposed daily outflows for the day {propFocDate}:
    {messageForEmail}
    Proposals Analyzed:

    {analyzedNames if len(analyzedNames) > 0 else 'No Proposals Analyzed'}


    Thank You.
    '''
        newEmail.To = email 
        newEmail.Send() 


    #Serveillance Database Program Implementation
    propFlaggedFiles_DB = SAPFileDetector(propFilenames, priorBusinessDay, 'Proposal') 
    regFlaggedFiles = SAPFileDetector(regFilenames, priorBusinessDay, 'Register') 
    treasFile = TreasuraFileDetector(treasDir, treasFilenames, currentDate) 
    finalData = DataAgreggator_DB(companyList, currencyList, accountList, propFlaggedFiles_DB, propDir, propRecvDate_DB, regFlaggedFiles, regDir, priorBusinessDay, treasFile)

    #Update the Database
    chmod(depositPath, S_IWUSR | S_IWOTH) 
    DATABASE_WORKBOOK = load_workbook(depositPath)
    dataSheet = DATABASE_WORKBOOK['Database']
    dataSheet.insert_rows(idx = 2, amount = 9)
    for j in range(finalData.shape[1]):

        #Built-in formats: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
        cellFormat = 'General'
        if j < 2:
            pass
        elif j < 4 or j == 15:
            cellFormat = 'mm-dd-yy'
        elif j < 6:
            pass
        else:
            cellFormat = r'$ #,###.00;[red]$ (#,###.00);$ 0.00;'

        for i in range(finalData.shape[0]):
            cell = dataSheet.cell(row = i + 3, column = j + 1)
            cell.value = finalData[i, j]
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = cellFormat

    DATABASE_WORKBOOK.save(depositPath)
    chmod(depositPath, S_IRUSR | S_IROTH) 

    chmod(dependenciesDir + '$$$$$$$$$$', S_IWUSR | S_IWOTH) 
    with open(dependenciesDir + '$$$$$$$$$$', 'a') as f:
        f.write(currentDate + ';') 
    chmod(dependenciesDir + '$$$$$$$$$$', S_IRUSR | S_IROTH) 

    print("Surveillance Database Updated Successfully")


#Program Execution
EXIT = False
try:
    UpdateInbox()
except Exception: 
    print_exc() 
    EXIT = True 

if not EXIT: 
    try:
        wd = getcwd()
        propDir = dependenciesDir + '$$$$$$$$$$'
        regDir = dependenciesDir + '$$$$$$$$$$'
        treasDir = dependenciesDir + '$$$$$$$$$$'
        propFocDate, currentDate, priorBusinessDay, propRecvDate_DB = InitializeApplication() 
        AcquireProgramLock() 
        AcquireFilesFromOutlook(priorBusinessDay, currentDate, regDir, treasDir) 
        ExecuteProposedOutflowsAndDatabaseUpdates(wd, propDir, regDir, treasDir, propFocDate, currentDate, priorBusinessDay, propRecvDate_DB) 
        PropRegTreasFolderCleanup(currentDate, propDir, regDir, treasDir) 
    except SystemExit:
        print('Program Terminated')
    except Exception:
        print_exc()
    finally:
        ReleaseProgramLock()
